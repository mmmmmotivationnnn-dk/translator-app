# --- START OF FINAL app.py (v1.6 - Multi-Model Support) ---
import os
import json
import re
import difflib
from urllib.parse import quote
import io
from flask import send_file
import openpyxl
from openpyxl.styles import Font
from flask_migrate import Migrate

# Новий імпорт для OpenRouter/OpenAI
from openai import OpenAI

import google.generativeai as genai
from flask import Flask, request, jsonify, render_template, abort, Response
from flask_sqlalchemy import SQLAlchemy

# --- НАЛАШТУВАННЯ ---
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'my-super-secret-key-for-translator-12345')

# --- НАЛАШТУВАННЯ БАЗИ ДАНИХ ---
# ... (код без змін) ...
if 'DATABASE_URL' in os.environ:
    # 1. Беремо URL з оточення
    db_url = os.environ['DATABASE_URL'].replace("postgres://", "postgresql://", 1)
    
    # 2. Додаємо параметр, який вимагає SSL, але з менш суворими налаштуваннями
    app.config['SQLALCHEMY_DATABASE_URI'] = f"{db_url}?sslmode=require"
else:
    DB_PASSWORD = '12345321'
    app.config['SQLALCHEMY_DATABASE_URI'] = f'postgresql://postgres:{DB_PASSWORD}@localhost/mytranslator_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# === ДОДАЙТЕ ЦІ ДВА РЯДКИ ПІСЛЯ ІНІЦІАЛІЗАЦІЇ `db` ===
migrate = Migrate(app, db)
# ====================================================

# --- НАЛАШТУВАННЯ API КЛІЄНТІВ ---
# Gemini
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', "AIzaSyCh9p83_R5EeCWQlU_kgsTShzzTCtqoQPQ")  # Вставте свій ключ Gemini
genai.configure(api_key=GEMINI_API_KEY)
gemini_model = genai.GenerativeModel('gemini-2.5-flash')

# OpenRouter (використовує OpenAI клієнт)
OPENROUTER_API_KEY = os.getenv('OPENROUTER_API_KEY',
                               "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")  # Вставте свій ключ OpenRouter
openrouter_client = OpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key=OPENROUTER_API_KEY,
)
# Ви можете змінювати модель тут
OPENROUTER_MODEL_NAME = "openai/gpt-oss-20b:free"


# --- МОДЕЛІ БАЗИ ДАНИХ ---
# ... (код без змін) ...
class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(100), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    parts = db.relationship('TextPart', backref='project', lazy=True, cascade="all, delete-orphan")


class TextPart(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    part_number = db.Column(db.Integer, nullable=False)
    part_title = db.Column(db.String(200), nullable=False)
    content_json = db.Column(db.Text, nullable=False, default='[]')


# --- ДОПОМІЖНІ ФУНКЦІЇ ---
def segment_with_separators(text: str) -> list[dict]:
    # ... (код без змін) ...
    if not text: return []
    lines = text.splitlines(keepends=True)
    grouped = []
    for line in lines:
        if not line.strip():
            grouped.append({'text': '', 'separator': line})
        else:
            sentence_parts = re.findall(r'.+?[.!?][\'"”)]*\s*|.+?$', line)
            for part in sentence_parts:
                if not part.strip(): continue
                match = re.search(r'(\s*)$', part)
                separator = match.group(1) if match else ""
                sentence_text = part[:len(part) - len(separator)]
                if sentence_text or separator: grouped.append({'text': sentence_text, 'separator': separator})
    return grouped


# --- НОВІ ФУНКЦІЇ ДЛЯ РОБОТИ З РІЗНИМИ МОДЕЛЯМИ ---
def get_translation_from_model(prompt: str, model_provider: str) -> str:
    """Викликає відповідну модель для перекладу."""
    try:
        if model_provider == 'openrouter':
            completion = openrouter_client.chat.completions.create(
                model=OPENROUTER_MODEL_NAME,
                messages=[{"role": "user", "content": prompt}]
            )
            return completion.choices[0].message.content.strip()
        else:  # Gemini за замовчуванням
            response = gemini_model.generate_content(prompt)
            return response.text.strip()
    except Exception as e:
        print(f"Error with model {model_provider}: {e}")
        return f"ERROR: Translation failed with {model_provider}."


def get_batch_translation_from_model(prompt: str, model_provider: str, sentence_count: int) -> list[str]:
    """Викликає відповідну модель для пакетного перекладу (повертає JSON)."""
    text_result = get_translation_from_model(prompt, model_provider)
    try:
        clean_text = re.sub(r'```json\s*|\s*```', '', text_result)
        translated_data = json.loads(clean_text)
        if not isinstance(translated_data, list) or len(translated_data) != sentence_count:
            raise ValueError("Mismatched sentence count.")
        return translated_data
    except (json.JSONDecodeError, ValueError) as e:
        print(f"CRITICAL PARSING ERROR with {model_provider}: {e}. Raw response: {text_result}")
        # В разі помилки, повертаємо список помилок, щоб не зламати фронтенд
        return ["Parsing Error"] * sentence_count


# --- МАРШРУТИ НАВІГАЦІЇ (без змін) ---
@app.route('/')
def show_categories():
    categories = [c[0] for c in db.session.query(Project.category).distinct().order_by(Project.category).all()]
    return render_template('index.html', categories=categories)


# ... (інші маршрути навігації) ...
@app.route('/category/<string:category_name>')
def show_projects_in_category(category_name):
    projects = Project.query.filter_by(category=category_name).order_by(Project.title).all()
    return render_template('category.html', projects=projects, category_name=category_name)


@app.route('/project/<int:project_id>')
def show_project_parts(project_id):
    project = Project.query.get_or_404(project_id)
    parts = sorted(project.parts, key=lambda x: x.part_number)
    return render_template('project_view.html', project=project, parts=parts)


# Знайдіть стару функцію @app.route('/project/<int/project_id>/export_xlsx')
# і повністю замініть її на цей виправлений код:

@app.route('/project/<int:project_id>/export_xlsx')
def export_project_xlsx(project_id):
    """
    Генерує та віддає проєкт у форматі .xlsx.
    Кожна частина проєкту - в окремому рядку колонки А.
    """
    project = Project.query.get_or_404(project_id)
    parts = sorted(project.parts, key=lambda x: x.part_number)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = project.title[:30]

    for index, part in enumerate(parts):
        full_english_text_for_part = ""
        try:
            content = json.loads(part.content_json)
            for item in content:
                # === ОСНОВНЕ ВИПРАВЛЕННЯ ТУТ ===
                # Замінюємо 'text' на 'en', щоб брати правильні дані
                full_english_text_for_part += item.get('en', '') + item.get('en_sep', ' ')
                # ===============================
        except (json.JSONDecodeError, TypeError):
            full_english_text_for_part = "Error: Could not decode content."

        ws.cell(row=index + 1, column=1, value=full_english_text_for_part.strip())

    ws.column_dimensions['A'].width = 100
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{project.title} (English).xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/translator/<int:part_id>', methods=['GET'])
def translator(part_id):
    part = TextPart.query.get_or_404(part_id)
    return render_template('translator.html', part=part)


# --- ОНОВЛЕНІ API МАРШРУТИ ---
@app.route('/api/translate_batch', methods=['POST'])
def translate_batch_api():
    data = request.json
    content_to_translate = data.get('content', [])
    model_provider = data.get('model_provider', 'gemini')
    if not content_to_translate: return jsonify({"translated_sentences": []})

    # === ФІНАЛЬНИЙ, НАЙБІЛЬШ НАДІЙНИЙ ПРОМПТ ДЛЯ ПЕРЕКЛАДУ ===
    # Ми створюємо JSON-структуру для входу, щоб модель точно зрозуміла завдання.
    input_json = json.dumps(
        [item.get('en', '') for item in content_to_translate],
        ensure_ascii=False
    )
    prompt = (
        f"Translate the English sentences in the following JSON array into Ukrainian.\n"
        f"Return a JSON array with the exact same number of elements, containing only the translated strings.\n"
        f"INPUT: {input_json}\n"
        f"OUTPUT:"
    )

    translated_data = get_batch_translation_from_model(prompt, model_provider, len(content_to_translate))
    return jsonify({"translated_sentences": translated_data})


@app.route('/api/part/<int:part_id>/sync', methods=['POST'])
def sync_and_save_part(part_id):
    part = TextPart.query.get_or_404(part_id)
    data = request.json
    new_uk_full_text = data.get('ukrainian_text', '')
    model_provider = data.get('model_provider', 'gemini')

    try:
        old_content_data = json.loads(part.content_json)
    except:
        old_content_data = []

    old_uk_sentences = [item.get('uk', '') for item in old_content_data]
    new_uk_parts = segment_with_separators(new_uk_full_text)
    new_uk_sentences = [p['text'] for p in new_uk_parts]

    matcher = difflib.SequenceMatcher(None, old_uk_sentences, new_uk_sentences, autojunk=False)
    new_content_json, next_id = [], max([item.get('id', 0) for item in old_content_data] + [0]) + 1

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            for i in range(i2 - i1):
                old_item = old_content_data[i1 + i]
                new_item = old_item.copy()
                new_item['uk_sep'] = new_uk_parts[j1 + i]['separator']
                new_item['en_sep'] = new_uk_parts[j1 + i]['separator']
                new_content_json.append(new_item)
        elif tag == 'insert':
            for k in range(j1, j2):
                new_uk_part = new_uk_parts[k]
                # === ФІНАЛЬНИЙ ПРОМПТ ДЛЯ РЕВЕРС-ПЕРЕКЛАДУ ===
                prompt = f"Translate to English: \"{new_uk_part['text']}\""
                new_en_text = get_translation_from_model(prompt, model_provider)
                new_content_json.append(
                    {'id': next_id, 'en': new_en_text, 'en_sep': new_uk_part['separator'], 'uk': new_uk_part['text'],
                     'uk_sep': new_uk_part['separator']})
                next_id += 1
        elif tag == 'replace':
            old_en_chunk = "".join([s.get('en', '') + s.get('en_sep', ' ') for s in old_content_data[i1:i2]]).rstrip()
            old_uk_chunk = "".join([s.get('uk', '') + s.get('uk_sep', ' ') for s in old_content_data[i1:i2]]).rstrip()
            new_uk_chunk = "".join([p['text'] + p['separator'] for p in new_uk_parts[j1:j2]]).rstrip()

            # === ФІНАЛЬНИЙ ПРОМПТ ДЛЯ СИНХРОНІЗАЦІЇ ===
            prompt = (
                f"The English text `{old_en_chunk}` was translated as `{old_uk_chunk}`. "
                f"A user corrected it to `{new_uk_chunk}`. "
                f"Update the English text to match the correction. "
                f"The output must be in English and have exactly {len(new_uk_parts[j1:j2])} sentences. "
                f"Return only the updated English text."
            )
            new_en_chunk = get_translation_from_model(prompt, model_provider)
            new_en_parts = segment_with_separators(new_en_chunk)

            len_en, len_uk = len(new_en_parts), j2 - j1
            if len_en > len_uk > 0:
                last = new_en_parts[len_uk - 1]
                for i in range(len_uk, len_en): last['text'] += new_en_parts[i]['separator'] + new_en_parts[i]['text']
                last['separator'] = new_en_parts[-1]['separator']
                new_en_parts = new_en_parts[:len_uk]

            for k in range(len(new_uk_parts[j1:j2])):
                en_text = new_en_parts[k]['text'] if k < len(new_en_parts) else ""
                en_sep = new_en_parts[k]['separator'] if k < len(new_en_parts) else " "
                uk_part = new_uk_parts[j1 + k]
                new_content_json.append({'id': next_id, 'en': en_text, 'en_sep': en_sep, 'uk': uk_part['text'],
                                         'uk_sep': uk_part['separator']})
                next_id += 1

    part.content_json = json.dumps(new_content_json, ensure_ascii=False)
    db.session.commit()
    return jsonify(new_content_json)


# Вставте цей код в app.py

@app.route('/api/part/<int:part_id>/sync_sentence', methods=['POST'])
def sync_sentence_api(part_id):
    """
    Надшвидкий API для синхронізації ОДНОГО речення.
    """
    part = TextPart.query.get_or_404(part_id)
    data = request.json

    sentence_id = data.get('id')
    new_uk_text = data.get('uk_text')
    model_provider = data.get('model_provider', 'gemini')

    try:
        content_data = json.loads(part.content_json)
    except:
        return jsonify({"error": "Failed to parse content"}), 500

    target_sentence = None
    for sentence in content_data:
        if sentence.get('id') == sentence_id:
            target_sentence = sentence
            break

    if not target_sentence:
        return jsonify({"error": "Sentence not found"}), 404

    old_en_text = target_sentence.get('en', '')
    old_uk_text = target_sentence.get('uk', '')

    # Використовуємо той самий надійний промпт для оновлення
    prompt = (
        f"The English text `{old_en_text}` was translated as `{old_uk_text}`. "
        f"A user corrected it to `{new_uk_text}`. "
        f"Update the English text to match the correction. "
        f"The output must be in English and be a single sentence. "
        f"Return only the updated English text."
    )
    new_en_text = get_translation_from_model(prompt, model_provider)

    # Оновлюємо дані в JSON і зберігаємо
    target_sentence['en'] = new_en_text
    target_sentence['uk'] = new_uk_text
    part.content_json = json.dumps(content_data, ensure_ascii=False)
    db.session.commit()

    return jsonify({"new_en_text": new_en_text})

@app.route('/add_project', methods=['POST'])
def add_project_api():
    auth_header = request.headers.get('Authorization')
    if not auth_header or auth_header != f"Bearer {app.config['SECRET_KEY']}":
        abort(401, description="Unauthorized")
    data = request.json
    new_project = Project(category=data['category'], title=data['title'])
    db.session.add(new_project)
    for part_data in data['parts']:
        content_with_seps = []
        for item in part_data['content_json']:
            item.setdefault('en_sep', ' ');
            item.setdefault('uk_sep', ' ')
            content_with_seps.append(item)
        new_part = TextPart(project=new_project, part_number=part_data['part_number'],
                            part_title=part_data['part_title'],
                            content_json=json.dumps(content_with_seps, ensure_ascii=False))
        db.session.add(new_part)
    db.session.commit()
    return jsonify({"message": f"Project '{new_project.title}' added successfully!", "id": new_project.id}), 201


@app.route('/api/categories')
def get_categories_api():
    categories = [c[0] for c in db.session.query(Project.category).distinct().order_by(Project.category).all()]
    return jsonify(categories)


@app.route('/project/<int:project_id>/export')
def export_project(project_id):
    project = Project.query.get_or_404(project_id)
    parts = sorted(project.parts, key=lambda x: x.part_number)
    full_english_text = ""
    for part in parts:
        full_english_text += f"--- Part {part.part_number}: {part.part_title} ---\n\n"
        try:
            content = json.loads(part.content_json)
            full_english_text += "".join([item.get('en', '') + item.get('en_sep', ' ') for item in content]) + "\n\n"
        except (json.JSONDecodeError, TypeError):
            full_english_text += "Error\n\n"
    encoded_filename = quote(f"{project.category} - {project.title}.txt")
    return Response(full_english_text.encode('utf-8'), mimetype="text/plain; charset=utf-8",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"})


with app.app_context(): db.create_all()
if __name__ == '__main__': app.run(debug=True)
# --- END OF FINAL app.py (v1.6 - Multi-Model Support) ---